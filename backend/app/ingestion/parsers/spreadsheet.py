import csv
import datetime as dt
import io
from decimal import Decimal
from typing import Never

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.ingestion.contracts import (
    FileKind,
    ParsedCell,
    ParsedDocument,
    ParsedRow,
    ParsedSheet,
    ValidatedFile,
    ValueKind,
)
from app.ingestion.errors import FileParsingError, IngestionErrorCode
from app.ingestion.limits import IngestionLimits
from app.ingestion.parsers.base import is_formula_like, normalize_text, safe_sheet_name


def _raise(code: IngestionErrorCode) -> Never:
    raise FileParsingError(code, "File parsing failed.")


def _cell_text(value: object, data_type: str) -> tuple[str, ValueKind, bool]:
    if data_type == "f":
        text = str(value)
        if not text.startswith("="):
            text = f"={text}"
        return normalize_text(text), ValueKind.FORMULA, True
    if data_type == "e":
        return normalize_text(str(value)), ValueKind.ERROR, False
    if isinstance(value, bool):
        return ("true" if value else "false"), ValueKind.BOOLEAN, False
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat(), ValueKind.DATE, False
    if isinstance(value, (int, float, Decimal)):
        return str(value), ValueKind.NUMBER, False
    text = normalize_text(str(value))
    return text, ValueKind.TEXT, is_formula_like(text)


def parse_xlsx(upload: ValidatedFile, limits: IngestionLimits) -> ParsedDocument:
    workbook = None
    try:
        workbook = load_workbook(
            io.BytesIO(upload.data),
            read_only=True,
            data_only=False,
            keep_links=False,
            keep_vba=False,
        )
        if len(workbook.worksheets) > limits.spreadsheet_sheets:
            _raise(IngestionErrorCode.SHEET_LIMIT_EXCEEDED)
        sheets: list[ParsedSheet] = []
        total_rows = 0
        total_cells = 0
        total_text = 0
        formula_like_count = 0
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            if (
                worksheet.max_row is not None
                and worksheet.max_row > limits.spreadsheet_rows_per_sheet
            ):
                _raise(IngestionErrorCode.ROW_LIMIT_EXCEEDED)
            if (
                worksheet.max_column is not None
                and worksheet.max_column > limits.spreadsheet_columns_per_row
            ):
                _raise(IngestionErrorCode.COLUMN_LIMIT_EXCEEDED)
            rows: list[ParsedRow] = []
            for row_number, source_row in enumerate(worksheet.iter_rows(), start=1):
                if row_number > limits.spreadsheet_rows_per_sheet:
                    _raise(IngestionErrorCode.ROW_LIMIT_EXCEEDED)
                if len(source_row) > limits.spreadsheet_columns_per_row:
                    _raise(IngestionErrorCode.COLUMN_LIMIT_EXCEEDED)
                cells: list[ParsedCell] = []
                for column_number, source_cell in enumerate(source_row, start=1):
                    if source_cell.value is None:
                        continue
                    value_text, value_kind, formula_like = _cell_text(
                        source_cell.value, source_cell.data_type
                    )
                    if len(value_text) > limits.spreadsheet_cell_characters:
                        _raise(IngestionErrorCode.TEXT_LIMIT_EXCEEDED)
                    total_cells += 1
                    total_text += len(value_text)
                    formula_like_count += int(formula_like)
                    if total_cells > limits.spreadsheet_total_cells:
                        _raise(IngestionErrorCode.CELL_LIMIT_EXCEEDED)
                    if total_text > limits.total_text_characters:
                        _raise(IngestionErrorCode.TEXT_LIMIT_EXCEEDED)
                    cells.append(
                        ParsedCell(
                            row_number=row_number,
                            column_number=column_number,
                            coordinate=f"{get_column_letter(column_number)}{row_number}",
                            value_text=value_text,
                            value_kind=value_kind,
                            formula_like=formula_like,
                        )
                    )
                if cells:
                    rows.append(ParsedRow(row_number=row_number, cells=tuple(cells)))
            total_rows += len(rows)
            sheets.append(
                ParsedSheet(
                    sheet_index=sheet_index,
                    name=safe_sheet_name(worksheet.title, f"Sheet {sheet_index}"),
                    rows=tuple(rows),
                )
            )
        warnings = ("FORMULA_LIKE_CELLS",) if formula_like_count else ()
        return ParsedDocument(
            kind=FileKind.XLSX,
            sheets=tuple(sheets),
            warnings=warnings,
            page_count=0,
            sheet_count=len(sheets),
            row_count=total_rows,
            cell_count=total_cells,
            text_length=total_text,
        )
    except FileParsingError:
        raise
    except (OSError, ValueError, TypeError, KeyError, IndexError, OverflowError):
        _raise(IngestionErrorCode.MALFORMED_FILE)
    finally:
        if workbook is not None:
            workbook.close()


def parse_csv(upload: ValidatedFile, limits: IngestionLimits) -> ParsedDocument:
    try:
        text = upload.data.decode("utf-8-sig")
        previous_limit = csv.field_size_limit()
        csv.field_size_limit(limits.spreadsheet_cell_characters)
        rows: list[ParsedRow] = []
        total_cells = 0
        total_text = 0
        formula_like_count = 0
        try:
            reader = csv.reader(io.StringIO(text, newline=""), dialect="excel", strict=True)
            for row_number, source_row in enumerate(reader, start=1):
                if row_number > limits.spreadsheet_rows_per_sheet:
                    _raise(IngestionErrorCode.ROW_LIMIT_EXCEEDED)
                if len(source_row) > limits.spreadsheet_columns_per_row:
                    _raise(IngestionErrorCode.COLUMN_LIMIT_EXCEEDED)
                cells: list[ParsedCell] = []
                for column_number, source_value in enumerate(source_row, start=1):
                    value = normalize_text(source_value)
                    formula_like = is_formula_like(value)
                    total_cells += 1
                    total_text += len(value)
                    formula_like_count += int(formula_like)
                    if total_cells > limits.spreadsheet_total_cells:
                        _raise(IngestionErrorCode.CELL_LIMIT_EXCEEDED)
                    if total_text > limits.total_text_characters:
                        _raise(IngestionErrorCode.TEXT_LIMIT_EXCEEDED)
                    cells.append(
                        ParsedCell(
                            row_number=row_number,
                            column_number=column_number,
                            coordinate=f"{get_column_letter(column_number)}{row_number}",
                            value_text=value,
                            value_kind=ValueKind.TEXT,
                            formula_like=formula_like,
                        )
                    )
                rows.append(ParsedRow(row_number=row_number, cells=tuple(cells)))
        finally:
            csv.field_size_limit(previous_limit)
        sheet = ParsedSheet(sheet_index=1, name="CSV", rows=tuple(rows))
        warnings = ("FORMULA_LIKE_CELLS",) if formula_like_count else ()
        return ParsedDocument(
            kind=FileKind.CSV,
            sheets=(sheet,),
            warnings=warnings,
            page_count=0,
            sheet_count=1,
            row_count=len(rows),
            cell_count=total_cells,
            text_length=total_text,
        )
    except FileParsingError:
        raise
    except (csv.Error, UnicodeError, ValueError, TypeError):
        _raise(IngestionErrorCode.MALFORMED_FILE)
